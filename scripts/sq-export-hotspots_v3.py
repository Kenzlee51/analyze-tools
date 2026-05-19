#!/usr/bin/env python3
"""
=============================================================================
sq-export-hotspots.py — Скрипт выгрузки Security Hotspots из SonarQube
=============================================================================

ОПИСАНИЕ:
    Читает БД проектов из SQ/sq-projects.json, выгружает Security Hotspots
    для каждого проекта со статусом success и сохраняет в xlsx.

    По умолчанию пропускает проекты у которых hotspots_status = "exported".
    Флаг --force принудительно перевыгружает все проекты.

ИСПОЛЬЗОВАНИЕ:
    python3 sq-export-hotspots.py [OPTIONS]

ОПЦИИ:
    --single-project NAME   Выгрузить только один указанный проект
    --force                 Принудительно перевыгрузить все проекты
                            (удаляет старые xlsx и скачивает заново)
    -h, --help              Показать справку

ПРИМЕРЫ:
    python3 sq-export-hotspots.py
    python3 sq-export-hotspots.py --single-project PROJ1
    python3 sq-export-hotspots.py --force
    python3 sq-export-hotspots.py --single-project PROJ1 --force

ОЖИДАЕМАЯ СТРУКТУРА:
    BASE_DIR/
    ├── scripts/
    │   └── sq-export-hotspots.py
    ├── SQ/
    │   └── sq-projects.json          ← БД проектов
    ├── logs/
    │   └── SQ-export/
    │       ├── run_YYYYMMDD_HHMMSS.log
    │       └── projects/
    │           └── PROJ1.log
    └── results/
        └── PROJ1/
            └── sq-hotspots/
                └── PROJ1_hotspots.xlsx

ЗАВИСИМОСТИ:
    Python 3.6+, requests, openpyxl
=============================================================================
"""

import os
import sys
import re
import json
import argparse
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# =============================================================================
# НАСТРАИВАЕМЫЕ ПАРАМЕТРЫ
# =============================================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SONARQUBE_URL   = "http://192.168.25.85:9000"
SONARQUBE_TOKEN = "squ_2ea978eff3573e01550db66c6959968b8cd498b4"

PAGE_SIZE = 500

RESULTS_DIR    = os.path.join(BASE_DIR, "results")
LOG_DIR        = os.path.join(BASE_DIR, "logs", "SQ-export")
SQ_DIR         = os.path.join(BASE_DIR, "SQ")
SQ_PROJECTS_DB = os.path.join(SQ_DIR, "sq-projects.json")
# =============================================================================

# =============================================================================
# ОЧИСТКА СТРОК ОТ СИМВОЛОВ НЕДОПУСТИМЫХ В XML/XLSX
# xlsx внутри — XML. Управляющие символы (кроме \t \n \r) вызывают
# "Ошибка в части содержимого книги" при открытии в Excel.
# =============================================================================
# Регулярка: все управляющие символы кроме TAB (x09), LF (x0A), CR (x0D)
_ILLEGAL_XML_CHARS = re.compile(
    r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]'
)

def sanitize(value):
    """
    Очищает строку от символов недопустимых в XML/xlsx.
    Возвращает строку или исходное значение если это не строка.
    """
    if not isinstance(value, str):
        return value
    # Убираем управляющие символы
    cleaned = _ILLEGAL_XML_CHARS.sub('', value)
    # Обрезаем до 32767 символов — максимум для ячейки Excel
    if len(cleaned) > 32767:
        cleaned = cleaned[:32764] + '...'
    return cleaned

def sanitize_row(row):
    """Очищает все значения в строке."""
    return [sanitize(cell) for cell in row]

# =============================================================================
# МАППИНГ securityCategory → читаемое название
# =============================================================================
CATEGORY_MAP = {
    "dos":                          "Denial of Service (DoS)",
    "sql-injection":                "SQL Injection",
    "xss":                          "Cross-Site Scripting (XSS)",
    "csrf":                         "Cross-Site Request Forgery (CSRF)",
    "ssrf":                         "Server-Side Request Forgery (SSRF)",
    "open-redirect":                "Open Redirect",
    "xpath-injection":              "XPath Injection",
    "log-injection":                "Log Injection",
    "ldap-injection":               "LDAP Injection",
    "command-injection":            "Command Injection",
    "path-traversal-injection":     "Path Traversal",
    "weak-cryptography":            "Weak Cryptography",
    "auth":                         "Authentication",
    "insecure-conf":                "Insecure Configuration",
    "file-manipulation":            "File Manipulation",
    "others":                       "Others",
    "object-injection":             "Object Injection",
    "http-response-splitting":      "HTTP Response Splitting",
    "encryption-of-sensitive-data": "Encryption of Sensitive Data",
    "traceability":                 "Traceability",
    "permission":                   "Permission",
    "buffer-overflow":              "Buffer Overflow",
}

EXTENSION_TO_LANGUAGE = {
    "js":    "JavaScript",
    "ts":    "TypeScript",
    "py":    "Python",
    "java":  "Java",
    "cs":    "C#",
    "cpp":   "C++",
    "cc":    "C++",
    "cxx":   "C++",
    "c":     "C",
    "h":     "C/C++",
    "hpp":   "C++",
    "php":   "PHP",
    "rb":    "Ruby",
    "go":    "Go",
    "kt":    "Kotlin",
    "swift": "Swift",
    "html":  "HTML",
    "xml":   "XML",
    "sh":    "Shell",
    "scala": "Scala",
    "rs":    "Rust",
}


def get_language(component):
    ext = component.split(".")[-1].lower() if "." in component else ""
    return EXTENSION_TO_LANGUAGE.get(ext, ext.upper() if ext else "Unknown")


# =============================================================================
# ЛОГИРОВАНИЕ
# =============================================================================
class Logger:
    def __init__(self, run_log, project_log=None):
        self.run_log     = run_log
        self.project_log = project_log

    def _write(self, message, log_path):
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(message + '\n')
        except Exception as e:
            print("[LOG ERROR] {}: {}".format(log_path, e))

    def info(self, message):
        line = "[{}] [INFO] {}".format(datetime.now().strftime('%H:%M:%S'), message)
        print(line)
        self._write(line, self.run_log)
        if self.project_log:
            self._write(line, self.project_log)

    def warn(self, message):
        line = "[{}] [WARN] {}".format(datetime.now().strftime('%H:%M:%S'), message)
        print(line)
        self._write(line, self.run_log)
        if self.project_log:
            self._write(line, self.project_log)

    def error(self, message):
        line = "[{}] [ERROR] {}".format(datetime.now().strftime('%H:%M:%S'), message)
        print(line)
        self._write(line, self.run_log)
        if self.project_log:
            self._write(line, self.project_log)

    def set_project_log(self, project_log):
        self.project_log = project_log


# =============================================================================
# БД ПРОЕКТОВ
# =============================================================================
def load_projects_db():
    if not os.path.exists(SQ_PROJECTS_DB):
        return {"projects": {}}
    try:
        with open(SQ_PROJECTS_DB, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print("[ERROR] Failed to load {}: {}".format(SQ_PROJECTS_DB, e))
        return {"projects": {}}


def save_projects_db(db):
    db["updated_at"] = datetime.now().isoformat()
    try:
        with open(SQ_PROJECTS_DB, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print("[ERROR] Failed to save {}: {}".format(SQ_PROJECTS_DB, e))


def update_hotspots_status(db, project_name, status):
    if project_name in db["projects"]:
        db["projects"][project_name]["hotspots_status"]     = status
        db["projects"][project_name]["hotspots_exported_at"] = (
            datetime.now().isoformat() if status == "exported" else ""
        )
        save_projects_db(db)


# =============================================================================
# SONARQUBE API
# =============================================================================
def get_session():
    session      = requests.Session()
    session.auth = (SONARQUBE_TOKEN, '')
    return session


def check_sonarqube_connection(log):
    log.info("Checking SonarQube connection: {}".format(SONARQUBE_URL))
    try:
        session  = get_session()
        response = session.get(
            "{}/api/server/version".format(SONARQUBE_URL),
            timeout=10
        )
        if response.status_code == 200:
            log.info("SonarQube available, version: {}".format(response.text.strip()))
            return True
        else:
            log.error("SonarQube returned HTTP {}".format(response.status_code))
            return False
    except requests.exceptions.ConnectionError:
        log.error("Cannot connect to SonarQube at {}".format(SONARQUBE_URL))
        return False
    except Exception as e:
        log.error("Connection error: {}".format(e))
        return False


def fetch_hotspots(project_key, log):
    session      = get_session()
    all_hotspots = []

    log.info("Fetching hotspots for project: {}".format(project_key))

    # SonarQube API по умолчанию возвращает только TO_REVIEW.
    # Запрашиваем оба статуса отдельно и объединяем.
    for status in ("TO_REVIEW", "REVIEWED"):
        page = 1
        log.info("  Fetching status={}...".format(status))

        while True:
            try:
                response = session.get(
                    "{}/api/hotspots/search".format(SONARQUBE_URL),
                    params={
                        "projectKey": project_key,
                        "status":     status,
                        "ps":         PAGE_SIZE,
                        "p":          page,
                    },
                    timeout=30
                )

                if response.status_code == 404:
                    log.error("Project not found in SonarQube: {}".format(project_key))
                    return None
                if response.status_code == 401:
                    log.error("Unauthorized — check token")
                    return None
                if response.status_code != 200:
                    log.error("API error HTTP {}: {}".format(
                        response.status_code, response.text[:200]))
                    return None

                data     = response.json()
                hotspots = data.get("hotspots", [])
                total    = data.get("paging", {}).get("total", 0)

                all_hotspots.extend(hotspots)
                log.info("    Page {}: +{} (status={}, total for status: {})".format(
                    page, len(hotspots), status, total))

                if page * PAGE_SIZE >= total:
                    break

                page += 1

            except requests.exceptions.Timeout:
                log.error("Timeout fetching hotspots page {} status={}".format(page, status))
                return None
            except Exception as e:
                log.error("Error fetching hotspots: {}".format(e))
                return None

    log.info("Total hotspots fetched: {}".format(len(all_hotspots)))
    return all_hotspots


# =============================================================================
# ФОРМИРОВАНИЕ XLSX
# =============================================================================
def build_xlsx(project_name, project_key, hotspots, output_path, log):
    log.info("Building xlsx: {}".format(output_path))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Security Hotspots"

    # --- Заголовки ---
    headers = [
        "Rule", "Message", "Category", "Priority", "Severity",
        "Language", "File", "Line", "Status", "Resolution", "Comments"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Данные ---
    for h in hotspots:
        component = h.get("component", "")

        file_path = component
        if component.startswith("{}:".format(project_key)):
            file_path = component[len(project_key) + 1:]

        probability = h.get("vulnerabilityProbability", "")

        if probability == "HIGH":
            severity = "CRITICAL"
        elif probability == "MEDIUM":
            severity = "MAJOR"
        else:
            severity = "MINOR"

        row = sanitize_row([
            h.get("ruleKey", ""),
            h.get("message", ""),
            CATEGORY_MAP.get(h.get("securityCategory", ""), h.get("securityCategory", "")),
            probability,
            severity,
            get_language(component),
            file_path,
            h.get("line", ""),
            h.get("status", ""),
            h.get("resolution", ""),
            "",  # Comments — пустое
        ])

        ws.append(row)

    # --- Ширина колонок ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 70
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 20

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    try:
        wb.save(output_path)
        log.info("xlsx saved: {}".format(output_path))
        return True
    except Exception as e:
        log.error("Failed to save xlsx: {}".format(e))
        return False


# =============================================================================
# ОБРАБОТКА ОДНОГО ПРОЕКТА
# =============================================================================
def process_project(project_name, project_info, db, args, log):
    projects_log_dir = os.path.join(LOG_DIR, "projects")
    os.makedirs(projects_log_dir, exist_ok=True)
    project_log_path = os.path.join(projects_log_dir, "{}.log".format(project_name))

    with open(project_log_path, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("Project: {}\n".format(project_name))
        f.write("Started: {}\n".format(datetime.now().isoformat()))
        f.write("=" * 60 + "\n\n")

    log.set_project_log(project_log_path)

    log.info("=" * 50)
    log.info("Processing project: {}".format(project_name))
    log.info("=" * 50)

    project_key     = project_info.get("project_key", "")
    hotspots_status = project_info.get("hotspots_status", "never")

    log.info("Project key    : {}".format(project_key))
    log.info("Hotspots status: {}".format(hotspots_status))

    if not args.force and hotspots_status == "exported":
        exported_at = project_info.get("hotspots_exported_at", "")
        log.info("Already exported at {}. Skipping (use --force to re-export)".format(
            exported_at))
        log.set_project_log(None)
        return True

    output_dir  = os.path.join(RESULTS_DIR, project_name, "sq-hotspots")
    output_file = os.path.join(output_dir, "{}_hotspots.xlsx".format(project_name))

    if args.force and os.path.exists(output_file):
        try:
            os.remove(output_file)
            log.info("Removed old xlsx (--force): {}".format(output_file))
        except Exception as e:
            log.warn("Failed to remove old xlsx: {}".format(e))

    os.makedirs(output_dir, exist_ok=True)

    hotspots = fetch_hotspots(project_key, log)

    if hotspots is None:
        log.error("Failed to fetch hotspots for: {}".format(project_name))
        update_hotspots_status(db, project_name, "failed")
        log.set_project_log(None)
        return False

    if len(hotspots) == 0:
        log.info("No hotspots found for project: {}".format(project_name))

    success = build_xlsx(project_name, project_key, hotspots, output_file, log)

    if success:
        update_hotspots_status(db, project_name, "exported")
        log.info("Done: {} hotspots -> {}".format(len(hotspots), output_file))
    else:
        update_hotspots_status(db, project_name, "failed")
        log.error("Failed to build xlsx for: {}".format(project_name))

    log.set_project_log(None)
    return success


# =============================================================================
# MAIN
# =============================================================================
def main():
    parser = argparse.ArgumentParser(
        description='Export Security Hotspots from SonarQube to xlsx'
    )
    parser.add_argument('--single-project', metavar='NAME',
                        help='Export only one project')
    parser.add_argument('--force', action='store_true',
                        help='Force re-export even if already exported')
    args = parser.parse_args()

    os.makedirs(LOG_DIR, exist_ok=True)
    os.makedirs(os.path.join(LOG_DIR, "projects"), exist_ok=True)

    run_log_path = os.path.join(
        LOG_DIR,
        "run_{}.log".format(datetime.now().strftime('%Y%m%d_%H%M%S'))
    )

    with open(run_log_path, 'w', encoding='utf-8') as f:
        f.write("SQ Export Hotspots run started: {}\n".format(datetime.now().isoformat()))
        f.write("BASE_DIR: {}\n".format(BASE_DIR))
        f.write("SONARQUBE_URL: {}\n".format(SONARQUBE_URL))
        f.write("=" * 60 + "\n\n")

    log = Logger(run_log_path)
    log.info("BASE_DIR: {}".format(BASE_DIR))
    log.info("SONARQUBE_URL: {}".format(SONARQUBE_URL))
    if args.force:
        log.info("Mode: --force (re-export all)")
    if args.single_project:
        log.info("Mode: --single-project {}".format(args.single_project))

    if not check_sonarqube_connection(log):
        log.error("Cannot connect to SonarQube. Aborting.")
        sys.exit(1)

    if not os.path.exists(SQ_PROJECTS_DB):
        log.error("Projects DB not found: {}".format(SQ_PROJECTS_DB))
        log.error("Run sq-analyze.py first to create the DB")
        sys.exit(1)

    db           = load_projects_db()
    all_projects = db.get("projects", {})
    log.info("Loaded projects DB: {} projects total".format(len(all_projects)))

    success_projects = {
        name: info for name, info in all_projects.items()
        if info.get("status") == "success"
    }
    log.info("Projects with status=success: {}".format(len(success_projects)))

    if not success_projects:
        log.warn("No projects with status=success found in DB")
        sys.exit(0)

    if args.single_project:
        if args.single_project not in success_projects:
            if args.single_project not in all_projects:
                log.error("Project not found in DB: {}".format(args.single_project))
            else:
                log.error("Project '{}' has status='{}', not 'success'".format(
                    args.single_project,
                    all_projects[args.single_project].get("status", "unknown")
                ))
            sys.exit(1)
        projects_to_process = {args.single_project: success_projects[args.single_project]}
    else:
        projects_to_process = success_projects

    log.info("Projects to export: {}".format(len(projects_to_process)))
    log.info("Projects: {}".format(', '.join(sorted(projects_to_process.keys()))))

    start_time = datetime.now()
    results    = {}

    for project_name, project_info in sorted(projects_to_process.items()):
        success = process_project(project_name, project_info, db, args, log)
        results[project_name] = success

    elapsed       = datetime.now() - start_time
    success_count = sum(1 for v in results.values() if v)
    fail_count    = len(results) - success_count
    skipped_count = sum(
        1 for name, info in projects_to_process.items()
        if not args.force and info.get("hotspots_status") == "exported"
    )

    log.info("")
    log.info("=" * 50)
    log.info("Export complete!")
    log.info("=" * 50)
    log.info("Total projects : {}".format(len(projects_to_process)))
    log.info("Successful     : {}".format(success_count))
    log.info("Skipped        : {}".format(skipped_count))
    log.info("Failed         : {}".format(fail_count))
    log.info("Time elapsed   : {}".format(elapsed))
    log.info("Run log        : {}".format(run_log_path))

    if fail_count > 0:
        log.info("")
        log.info("Failed projects:")
        for name, ok in results.items():
            if not ok:
                log.error("  - {}".format(name))

    sys.exit(0 if fail_count == 0 else 1)


if __name__ == '__main__':
    main()
